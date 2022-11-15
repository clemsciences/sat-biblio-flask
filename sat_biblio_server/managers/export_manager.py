"""


"""

import codecs
import csv
import datetime
import logging
import os

import openpyxl
from openpyxl.styles import Font, colors, PatternFill, Border, Side, Protection, Alignment
from openpyxl.utils import get_column_letter

from sat_biblio_server.managers.catalogue_manager import CatalogueHamelain1, CatalogueHamelain3
from sat_biblio_server.data.models import Enregistrement


class ExportCatalogueManager:

    @staticmethod
    def export_catalogue_csv(path, filename_without_extension, records_db):
        records = [Enregistrement.from_db_to_data(record_db) for record_db in records_db]
        print(os.getcwd())
        complete_path = os.path.join(path, f"{filename_without_extension}.csv")
        with codecs.open(complete_path, "w", encoding="utf-8") as f:
            writer = csv.writer(f, delimiter="\t")
            writer.writerow(["Auteurs", "Titre", "Lieu d'édition", "Editeur", "Année", "Nombre de pages",
                             "Année d'obtention", "Description", "Commentaire", "Cote",
                             "Nombre d'exemplaire supplémentaire", "Provenance", "Mots-clés", "Ligne"])
            for record in records:
                # print(record)
                writer.writerow(Enregistrement.from_data_to_csv_row(record))
        return complete_path

    @staticmethod
    def export_catalogue_xlsx(path, filename_without_extension, records_db):
        complete_path = os.path.join(path, f"{filename_without_extension}.xlsx")
        excel_catalogue = openpyxl.Workbook()
        current_sheet = excel_catalogue.active
        current_sheet.title = "Catalogue"
        font = Font(bold=True)
        fill = PatternFill(start_color='FFFFFFFF', end_color='FF000000')
        border = Border(left=Side(color="FF000000"),
                        right=Side(color="FF000000"),
                        bottom=Side(color="FF000000"),
                        top=Side(color="FF000000"))
        protection = Protection(locked=True)
        column_names = Enregistrement.get_column_names()
        # column_names = ["Auteurs", "Titre", "Lieu d'édition", "Editeur", "Année", "Nombre de pages",
        #  "Année d'obtention", "Description", "Commentaire", "Cote",
        #  "Nombre d'exemplaire supplémentaire", "Provenance", "Mots-clés"]
        # region column widths
        column_widths = [30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30]
        for i, width in enumerate(column_widths, 1):
            column_letter = get_column_letter(i)
            current_sheet.column_dimensions[column_letter].width = width
            # current_sheet.column_dimensions[column_letter].height = 10
        current_sheet.auto_filter.ref = f"A1:N{len(records_db)+1}"
        current_sheet.auto_filter.add_sort_condition(f"{get_column_letter(10)}1:{get_column_letter(10)}{len(records_db)+1}")

        # endregion
        for i in range(len(column_names)):
            c = current_sheet.cell(row=1, column=i+1, value=column_names[i])
            c.font = font
            c.fill = fill
            c.border = border

        for i, record_db in enumerate(records_db):
            record = Enregistrement.from_db_to_data(record_db)
            row = Enregistrement.from_data_to_csv_row(record)
            current_sheet.row_dimensions[i].height = 30
            if i % 1000 == 0:
                logging.error(row)
            for j in range(len(row)):
                c = current_sheet.cell(row=i+2, column=j+1, value=row[j])
                c.alignment = Alignment(wrap_text=True)
        current_sheet.protection = protection

        excel_catalogue.save(complete_path)
        return complete_path

    @staticmethod
    def export_hamelain_1(path: str, catalogue: CatalogueHamelain1):
        complete_path = os.path.join(path, f"catalogue-exporte-hamelain-1.xlsx")
        excel_catalogue = openpyxl.Workbook()
        current_sheet = excel_catalogue.active
        current_sheet.title = "Catalogue"
        font = Font(bold=True)
        fill = PatternFill(start_color='FFFFFFFF', end_color='FF000000')
        border = Border(left=Side(color="FF000000"),
                        right=Side(color="FF000000"),
                        bottom=Side(color="FF000000"),
                        top=Side(color="FF000000"))
        protection = Protection(locked=True)
        column_names = CatalogueHamelain1.COLUMNS
        column_widths = [30, 30, 30, 30, 30, 30, 30, 30, 30, 30]
        for i, width in enumerate(column_widths, 1):
            column_letter = get_column_letter(i)
            current_sheet.column_dimensions[column_letter].width = width
            # current_sheet.column_dimensions[column_letter].height = 10
        current_sheet.auto_filter.ref = f"A1:N{len(catalogue.rows) + 1}"
        current_sheet.auto_filter.add_sort_condition(
            f"{get_column_letter(10)}1:{get_column_letter(10)}{len(catalogue.rows) + 1}")

        # endregion
        for i in range(len(column_names)):
            c = current_sheet.cell(row=1, column=i + 1, value=column_names[i])
            c.font = font
            c.fill = fill
            c.border = border

        for i, row_hamelain in enumerate(catalogue.rows):
            row = row_hamelain.to_csv_row()
            current_sheet.row_dimensions[i].height = 30
            # print(row)
            if i % 1000 == 0:
                logging.error(row)
            for j in range(len(row)):
                c = current_sheet.cell(row=i + 2, column=j + 1, value=row[j])
                c.alignment = Alignment(wrap_text=True)
        current_sheet.protection = protection

        excel_catalogue.save(complete_path)
        return complete_path

    @staticmethod
    def export_hamelain_3(path: str, ch3: CatalogueHamelain3, filename_prefix: str = "", title=""):
        complete_path = os.path.join(path, f"{filename_prefix}catalogue-exporte-hamelain-3.xlsx")
        excel_catalogue = openpyxl.Workbook()
        first_line = "Bibliothèque de la S.A.T. Inventaire des ouvrages. Format 2022. Point au 15-10-2022 par P.H."
        current_sheet = excel_catalogue.active
        current_sheet.title = "Catalogue"
        bold_font = Font(bold=True)
        white_fill = PatternFill(start_color='FFFFFFFF', end_color='FF000000')
        black_border = Border(left=Side(color="FF000000"),
                        right=Side(color="FF000000"),
                        bottom=Side(color="FF000000"),
                        top=Side(color="FF000000"))
        protection = Protection(locked=True)
        center_alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        column_names = CatalogueHamelain3.COLUMNS
        column_widths = CatalogueHamelain3.COLUMNS_WIDTHS

        shift_i_row = 1
        # region first line
        current_date = datetime.date.today().isoformat()
        c = current_sheet.cell(row=shift_i_row, column=1, value=f"BIBLIOTHEQUE DE LA SAT. Inventaire des livres. {title}. Point au {current_date}")
        c.font = bold_font
        c.fill = white_fill
        c.border = black_border
        c.alignment = center_alignment

        current_sheet.merge_cells(range_string=f"A1:{get_column_letter(len(CatalogueHamelain3.COLUMNS))}1")
        # endregion

        # region second line
        shift_i_row += 1

        for i in range(len(column_names)):
            c = current_sheet.cell(row=shift_i_row, column=i + 1, value=column_names[i])
            c.font = bold_font
            c.fill = white_fill
            c.border = black_border
            c.alignment = center_alignment


        for i, width in enumerate(column_widths, 1):
            column_letter = get_column_letter(i)
            current_sheet.column_dimensions[column_letter].width = width
            # current_sheet.column_dimensions[column_letter].height = 10
        current_sheet.auto_filter.ref = f"A{shift_i_row}:L{len(ch3.rows) + shift_i_row}"
        current_sheet.auto_filter.add_sort_condition(
            f"{get_column_letter(1)}{shift_i_row}:{get_column_letter(1)}{len(ch3.rows) + shift_i_row}")
        # endregion

        shift_i_row += 1

        for i, row_hamelain in enumerate(ch3.rows):
            row = row_hamelain.to_csv_row()
            current_sheet.row_dimensions[i].height = 30
            # print(row)
            if i % 1000 == 0:
                logging.error(row)
            for j in range(len(row)):
                c = current_sheet.cell(row=i + shift_i_row, column=j + 1, value=row[j])
                c.alignment = center_alignment
        current_sheet.protection = protection

        excel_catalogue.save(complete_path)
        return complete_path

