import datetime
import os
import random
from typing import Optional, List, Dict

import openpyxl
from werkzeug.utils import secure_filename

from config.production import Config
from sat_biblio_server.managers.catalogue_manager import CatalogueSchweitz, CatalogueHamelain1, \
    CatalogueHamelain2, CatalogueHamelain3, Catalogue2023
from sat_biblio_server import db, Author2023DB, ReferenceBibliographiqueLivre2023DB, Enregistrement2023DB, PACKDIR
from sat_biblio_server.data.models_2023 import Author2023, ReferenceBibliographiqueLivre2023, Enregistrement2023


class CsvReader:
    @staticmethod
    def read_csv(filename: str):
        pass

    @classmethod
    def get_current_sheet(cls, filename):
        work_book = openpyxl.load_workbook(filename=filename)
        return work_book.active


    @classmethod
    def read_xlsx(cls, filename: str, ignore_n_first_lines=0):
        current_sheet = cls.get_current_sheet(filename)
        rows = [row for i, row in enumerate(current_sheet.iter_rows(values_only=True)) if
                i > ignore_n_first_lines and row[0] is not None and row[1] is not None]
        return rows

    @classmethod
    def generate_xlsx(cls, columns: List[str]):
        excel_catalogue = openpyxl.Workbook(True)
        current_sheet = excel_catalogue.active
        current_sheet.title = "Catalogue"


    @classmethod
    def get_row(cls, filename: str, index: int):
        """

        :param filename:
        :param index:
        :return:
        """
        current_sheet = cls.get_current_sheet(filename)
        for i, row in enumerate(current_sheet.iter_rows(values_only=True)):
            if i == index:
                return row
        return []

    @classmethod
    def read_xslx_with_columns(cls, filename):
        current_sheet = cls.get_current_sheet(filename)
        # rows = [row for i, row in enumerate(current_sheet.iter_rows(values_only=True))
        #         if row[0] is not None and row[1] is not None]
        rows = [row for i, row in enumerate(current_sheet.iter_rows(values_only=True))
                if row[0] is not None and row[1] is not None]
        return rows

    @classmethod
    def read_first_lines(cls, filename: str, ignore_n_first_lines: int, n_first_lines: int):
        current_sheet = cls.get_current_sheet(filename)
        rows = [row for i, row in enumerate(current_sheet.iter_rows(values_only=True))
                if ignore_n_first_lines < i <= ignore_n_first_lines + n_first_lines]
        # print(rows)
        return rows



class ImportManager(CsvReader):
    @classmethod
    def import_schweitz_format(_cls, filename: str, ignore_n_first_lines=0):
        rows = _cls.read_xlsx(filename)
        catalogue = CatalogueSchweitz()
        catalogue.parse_rows(rows)
        return catalogue

    @classmethod
    def import_hamelain_1(_cls, filename: str, ignore_n_first_lines=0):
        rows = _cls.read_xlsx(filename)
        catalogue = CatalogueHamelain1()
        catalogue.parse_rows(rows, ignore_n_first_lines)
        return catalogue

        # excel_catalogue = openpyxl.load_workbook(filename)

    @classmethod
    def import_hamelain_2(_cls, filename: str, ignore_n_first_lines=0):
        # excel_catalogue = openpyxl.load_workbook(filename)
        rows = _cls.read_xlsx(filename)
        catalogue = CatalogueHamelain2()
        catalogue.parse_rows(rows, ignore_n_first_lines)
        return catalogue

    @classmethod
    def import_hamelain_3(_cls, filename: str, ignore_n_first_lines=0):
        rows = _cls.read_xlsx(filename)
        catalogue = CatalogueHamelain3()
        catalogue.parse_rows(rows, ignore_n_first_lines)
        return catalogue

    @classmethod
    def check_hamelain_3_column_names(cls, filename: str):
        rows = cls.read_xlsx(filename)
        return CatalogueHamelain3.check_column_names(rows)


class ImportManager2023(CsvReader):
    @classmethod
    def parse_file_to_catalogue_2023(cls, filename: str, ignore_n_first_lines=0):
        rows = cls.read_xslx_with_columns(filename)
        catalogue = Catalogue2023()
        catalogue.parse_rows(rows, ignore_n_first_lines)
        return catalogue

    @classmethod
    def check_2023_column_names(cls, filename: str, ignore_n_first_rows=0):
        catalogue = cls.parse_file_to_catalogue_2023(filename, ignore_n_first_rows)
        return catalogue.check_column_names()

    @classmethod
    def get_rows(cls, filename: str):
        catalogue = cls.parse_file_to_catalogue_2023(filename)
        return catalogue.rows_filtered_by_verified_entry

    @classmethod
    def import_catalogue_to_db(cls, catalogue: Catalogue2023):
        for i, row in enumerate(catalogue.rows_filtered_by_verified_entry):
            authors_db = []
            for author in row.authors:
                author_exists = Author2023DB.query.filter_by(first_name=author.first_name,
                                                         family_name=author.family_name).first()
                if author_exists:
                    author_db = author_exists
                else:
                    author_db = Author2023DB()
                    # print(repr(author.first_name))
                    # print(repr(author.family_name))
                    author_db.first_name = author.first_name
                    author_db.family_name = author.family_name
                    author_db.valide = True
                    author_db.origin = "import"
                    db.session.add(author_db)
                    db.session.commit()

                authors_db.append(author_db)

            ref = row.book_reference

            ref.authors = authors_db
            ref.origin = "import"
            db.session.add(ref)
            db.session.commit()

            rec = row.book_record
            rec.id_reference = ref.id
            rec.origin = "import"
            db.session.add(rec)
            db.session.commit()
            # print(rec.)
            # print(rec)
            # if i > 5:
            #     break


            # reference_db = ReferenceBibliographiqueLivre.from_data_to_db(ref)
            # # print(reference_db)
            #
            # record = icu.extraire_enregistrements(processed_row)
            # if record is None:
            #     print("record is None")
            #     continue
            # # print(record)
            # # region reference
            # db.session.add(reference_db)
            # db.session.commit()
            # # endregion
            # # region record
            # enregistrement_db = Enregistrement.from_data_to_db(record)
            # # print(enregistrement_db)
            # enregistrement_db.reference = reference_db
            # db.session.add(enregistrement_db)
            # db.session.commit()


class CatalogueFile:
    def __init__(self, timestamp: datetime.datetime, key: str, filename: str):
        self.timestamp = timestamp
        self.key = key
        self.filename = filename
    def to_dict(self) -> Optional[Dict[str, str]]:
        return dict(datetime=self.timestamp.isoformat(),
             key=self.key,
             filename=self.filename)

    @classmethod
    def from_dict_values(cls,
                         timestamp_string: str,
                         key: str,
                         filename: str):
        return cls(datetime.datetime.fromisoformat(timestamp_string),
                   key, filename)


class CatalogueFileManager:

    UPLOAD_FOLDER = os.path.join(PACKDIR, "..", Config.UPLOAD_FOLDER, Config.UPLOAD_CATALOGUE_FOLDER)

    @classmethod
    def generate_key(cls) -> str:
        """
        Generate a key.
        
        >>> random.seed(1994)
        >>> CatalogueFileManager.generate_key()
        'qpvjxolufa'

        :return: a 10-length string
        """
        id_size = 10
        key = "".join([random.choice("abcdefghijklmopqrstuvwxyz") for _ in range(id_size)])
        return key

    @classmethod
    def get_existing_keys(cls) -> List[str]:
        """
        Make 
        >>> CatalogueFileManager.get_existing_keys()

        :return:
        """
        keys = []
        if os.path.exists(cls.UPLOAD_FOLDER):
            for filename in os.listdir(cls.UPLOAD_FOLDER):
                parts = filename.split("_")
                if len(parts) > 2:
                    keys.append(parts[1])
        return keys

    @classmethod
    def get_catalogue_filename(cls, key: str) -> str:
        if os.path.exists(cls.UPLOAD_FOLDER):
            for filename in os.listdir(cls.UPLOAD_FOLDER):
                parts = filename.split("_")
                if len(parts) > 2 and parts[1] == key:
                    return filename
        return ""

    @classmethod
    def get_catalogue_path(cls, key: str) -> str:
        filename = cls.get_catalogue_filename(key)
        return os.path.join(cls.UPLOAD_FOLDER, filename)

    @classmethod
    def load_catalogue(cls, key: str):
        filename = cls.get_catalogue_filename(key)
        if filename:
            with open(os.path.join(cls.UPLOAD_FOLDER, filename), "rb") as f:
                return f.read()
        return None

    @classmethod
    def get_catalogue_count(cls) -> int:
        return len(cls.get_existing_keys())

    @classmethod
    def extract_from_filename(cls, filename: str) -> Optional[CatalogueFile]:
        parts = filename.split("_")
        timestamp_string = parts[0].replace("!", ":")
        # print(timestamp_string)
        # print(datetime.datetime.fromisoformat(timestamp_string))
        timestamp_datetime = datetime.datetime.fromisoformat(timestamp_string)
        if len(parts) == 3:
            catalogue_file = CatalogueFile(timestamp_datetime, parts[1], parts[2])
            return catalogue_file
            # dict(datetime=timestamp_string,
            #             key=parts[1],
            #             filename=parts[2])
        return None

    @classmethod
    def get_catalogue_list(cls, n_page: int, size: int) -> List[CatalogueFile]:
        """
        >>>
        >>> os.chdir("..")
        >>> CatalogueFileManager.get_catalogue_list(1, 3)

        >>> CatalogueFileManager.get_catalogue_list(2, 3)

        >>> CatalogueFileManager.get_catalogue_list(3, 3)

        >>> CatalogueFileManager.get_catalogue_list(4, 3)

        :param n_page: starts from 1
        :param size:
        :return:
        """
        page = []
        if os.path.exists(cls.UPLOAD_FOLDER):
            sorted_filenames = sorted(os.listdir(cls.UPLOAD_FOLDER))
            for i, filename in enumerate(sorted_filenames):
                current_page = i // size + 1
                if current_page == n_page:
                    page.append(cls.extract_from_filename(filename))
        return page

    @classmethod
    def save_downloaded_file(cls, file) -> Optional[str]:
        """

        :param file:
        :return: key of catalogue if file has been saved
        """
        # region new filename generation
        # region key generation
        key = cls.generate_key()
        existing_keys = cls.get_existing_keys()
        # ensure key unicity
        while key in existing_keys:
            key = cls.generate_key()
        # endregion
        filename = f"{datetime.datetime.now().isoformat().replace(':', '!')}" \
                   f"_{key}" \
                   f"_{secure_filename(file.filename).replace('_', '-')}"
        # endregion
        if not os.path.exists(cls.UPLOAD_FOLDER):
            os.makedirs(cls.UPLOAD_FOLDER)
        if os.path.exists(cls.UPLOAD_FOLDER):
            # file.save()

            path_to_file = os.path.join(cls.UPLOAD_FOLDER, filename)
            # with open(path_to_file, "wb") as f:
            #     f.write(file.)
            file.save(path_to_file)
            return key
        return None

    @classmethod
    def delete_catalogue(cls, key: str) -> bool:
        filename = cls.get_catalogue_filename(key)
        path_to_file = os.path.join(cls.UPLOAD_FOLDER, filename)
        if os.path.exists(path_to_file):
            os.remove(path_to_file)
            return True
        return False

    @classmethod
    def get_first_row(cls, key: str, ignore_n_first_rows: int, n_first_lines: int):
        path_to_file = cls.get_catalogue_path(key)
        rows = CsvReader.read_first_lines(path_to_file, ignore_n_first_rows, n_first_lines)
        return rows


