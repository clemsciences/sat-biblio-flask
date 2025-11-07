"""
Route to handle SAT subscription form submission: generates an XLSX attachment and sends emails.
"""
from io import BytesIO
from datetime import datetime

from flask import request
from flask_mail import Message

from sat_biblio_server import sat_biblio
from sat_biblio_server.utils import json_result
from sat_biblio_server.managers.mail_manager import send_email

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover
    Workbook = None
    Alignment = None
    get_column_letter = None

SAT_EMAIL = "admin@societearcheotouraine.fr"


def _compute_total(membership_type: str, receive_bulletin_by_post: bool) -> int:
    base = 0
    if membership_type == "student":
        base = 10
    elif membership_type == "single":
        base = 46
    elif membership_type == "couple":
        base = 62
    # if no membership selected we consider total 0
    if receive_bulletin_by_post:
        base += 13
    return base


PAYMENT_INSTRUCTIONS = {
    "cheque": (
        "Chèque",
        "Merci d'établir un chèque à l'ordre de la Société archéologique de Touraine.\n"
        "Adresse d'envoi: 37 avenue André Malraux.\n"
        "Vous pouvez aussi déposer le chèque à la BHT."
    ),
    "virement": (
        "Virement bancaire",
        "Merci d'effectuer un virement bancaire au compte de la SAT.\n"
        "IBAN: FR76 3000 0000 0000 0000 0000 000\nBIC: XXXXXX\n"
        "Indiquez en référence: Adhésion SAT - {last_name} {first_name}."
    ),
    "carte": (
        "Carte bancaire ou espèces",
        "Paiement par carte bancaire ou en espèces à la BHT (mardi 18h-21h ; mercredi et samedi 10h-13h)\n"
        "ou lors des séances mensuelles."
    )
}


def _build_payment_instruction(method: str, first_name: str, last_name: str) -> str:
    title, text = PAYMENT_INSTRUCTIONS.get(method, ("", ""))
    return f"Mode de règlement choisi: {title}\n\n{text.format(first_name=first_name, last_name=last_name)}"


def _build_xlsx(payload: dict) -> bytes:
    if Workbook is None:
        # fallback: simple CSV-like bytes
        headers = [
            "Prénom", "Nom", "Adresse", "Téléphone fixe", "Téléphone mobile", "Courriel",
            "Profession", "Année de naissance",
            "Prénom (2)", "Nom (2)", "Profession (2)", "Année de naissance (2)",
            "Type d'adhésion", "Bulletin par la poste", "Montant total", "Règlement",
            "Recevoir infos par courrier", "Refus publication"
        ]
        row = [
            payload.get("firstName", ""),
            payload.get("lastName", ""),
            payload.get("address", ""),
            payload.get("phoneLandline", ""),
            payload.get("phoneMobile", ""),
            payload.get("email", ""),
            payload.get("profession", ""),
            payload.get("birthYear", ""),
            payload.get("secondFirstName", ""),
            payload.get("secondLastName", ""),
            payload.get("secondProfession", ""),
            payload.get("secondBirthYear", ""),
            payload.get("membershipType", ""),
            "oui" if payload.get("receiveBulletinByPost") else "non",
            payload.get("totalAmount", 0),
            payload.get("paymentMethod", ""),
            "oui" if payload.get("mustReceiveInfoByMail") else "non",
            "oui" if payload.get("refusePublishInfo") else "non",
        ]
        content = ";".join(map(str, headers)) + "\n" + ";".join(map(lambda x: str(x).replace("\n", " "), row))
        return content.encode("utf-8")

    wb = Workbook()
    ws = wb.active
    ws.title = "Adhésion SAT"
    ws.append([
        "Prénom", "Nom", "Adresse", "Téléphone fixe", "Téléphone mobile", "Courriel",
        "Profession", "Année de naissance",
        "Prénom (2)", "Nom (2)", "Profession (2)", "Année de naissance (2)",
        "Type d'adhésion", "Bulletin par la poste", "Montant total", "Règlement",
        "Recevoir infos par courrier", "Refus publication", "Date"
    ])
    ws.append([
        payload.get("firstName", ""),
        payload.get("lastName", ""),
        payload.get("address", ""),
        payload.get("phoneLandline", ""),
        payload.get("phoneMobile", ""),
        payload.get("email", ""),
        payload.get("profession", ""),
        payload.get("birthYear", ""),
        payload.get("secondFirstName", ""),
        payload.get("secondLastName", ""),
        payload.get("secondProfession", ""),
        payload.get("secondBirthYear", ""),
        payload.get("membershipType", ""),
        "oui" if payload.get("receiveBulletinByPost") else "non",
        payload.get("totalAmount", 0),
        payload.get("paymentMethod", ""),
        "oui" if payload.get("mustReceiveInfoByMail") else "non",
        "oui" if payload.get("refusePublishInfo") else "non",
        datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    ])

    # Center content and wrap text, then auto-size columns and rows
    if Alignment is not None:
        # 1) Set alignment (center + wrap)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 2) Compute and apply column widths based on content length (max per column)
        if get_column_letter is not None:
            max_lengths = {}
            for col_idx in range(1, ws.max_column + 1):
                max_lengths[col_idx] = 0
                for row_idx in range(1, ws.max_row + 1):
                    value = ws.cell(row=row_idx, column=col_idx).value
                    if value is None:
                        continue
                    # handle multi-line: take the longest visual line
                    lines = str(value).splitlines() or [""]
                    longest = max(len(line) for line in lines)
                    if longest > max_lengths[col_idx]:
                        max_lengths[col_idx] = longest
            for col_idx, length in max_lengths.items():
                # Approximate: each character ~ 1 unit, add padding
                width = min(max(length + 2, 10), 60)
                ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 3) Adjust row heights based on number of lines in cells
        base_height = 15  # approx default row height in points
        for row_idx in range(1, ws.max_row + 1):
            max_lines = 1
            for col_idx in range(1, ws.max_column + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is None:
                    continue
                lines = str(value).splitlines() or [""]
                if len(lines) > max_lines:
                    max_lines = len(lines)
            # cap the row height to avoid excessively large rows
            ws.row_dimensions[row_idx].height = min(base_height * max_lines, 90)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _compose_email_subject(payload: dict) -> str:
    first_name = payload.get("firstName", "")
    last_name = payload.get("lastName", "")
    membership = payload.get("membershipType", "?")
    return f"Nouvelle adhésion SAT - {last_name} {first_name} ({membership})"


def _compose_email_body(payload: dict, to_member: bool = False) -> str:
    first_name = payload.get("firstName", "")
    last_name = payload.get("lastName", "")
    amount = payload.get("totalAmount", 0)
    method = payload.get("paymentMethod", "")
    instr = _build_payment_instruction(method, first_name, last_name)

    intro = (
        "Bonjour,\n\n"
        "Nous avons bien reçu votre demande d'adhésion à la Société Archéologique de Touraine.\n"
        if to_member else
        "Bonjour,\n\nUne nouvelle demande d'adhésion a été soumise.\n"
    )

    txt = (
        f"Adhérent: {first_name} {last_name}\n"
        "Tarifs valables du 01/01/2025 au 31/12/2025.\n"
        f"Montant à régler: {amount} €\n"
        f"{instr}\n\n"
        "Les informations fournies sont jointes à ce message sous forme de fichier .xlsx.\n\n"
        "Ceci est un message automatique."
    )
    return intro + txt


@sat_biblio.route("/send-sat", methods=["POST"])  # path required in issue description
@sat_biblio.route("/send-sat/", methods=["POST"])  # tolerate trailing slash
def send_sat_subscription():
    data = request.get_json() or {}

    # Basic required fields
    first_name = data.get("firstName", "").strip()
    last_name = data.get("lastName", "").strip()
    address = data.get("address", "").strip()
    membership_type = data.get("membershipType")
    receive_bulletin = bool(data.get("receiveBulletinByPost", False))
    payment_method = data.get("paymentMethod")

    if not first_name or not last_name or not address or not membership_type or not payment_method:
        return json_result(False, message="Champs manquants"), 400

    # compute total if not provided
    total = data.get("totalAmount")
    if total is None:
        total = _compute_total(membership_type, receive_bulletin)
        data["totalAmount"] = total

    # Build attachment
    xlsx_bytes = _build_xlsx(data)

    # Compose and send to SAT
    subject = _compose_email_subject(data)
    msg_admin = Message(subject=subject)
    msg_admin.body = _compose_email_body(data, to_member=False)
    msg_admin.add_recipient(SAT_EMAIL)
    filename = f"adhesion_sat_{last_name}_{first_name}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    msg_admin.attach(filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", xlsx_bytes)
    ok1 = send_email(msg_admin, SAT_EMAIL)

    # Send to member if email provided
    ok2 = True
    member_email = (data.get("email") or "").strip()
    if member_email:
        msg_member = Message(subject="Votre demande d'adhésion à la SAT")
        msg_member.body = _compose_email_body(data, to_member=True)
        msg_member.add_recipient(member_email)
        # msg_member.attach(filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", xlsx_bytes)
        ok2 = send_email(msg_member, member_email)

    if ok1 and ok2:
        return json_result(True, message="Demande d'adhésion envoyée."), 200
    else:
        return json_result(False, message="Echec de l'envoi d'email."), 500
